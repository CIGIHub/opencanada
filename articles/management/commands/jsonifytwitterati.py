import json
from operator import itemgetter

from django.core.management.base import BaseCommand
from openpyxl import load_workbook


class Command(BaseCommand):
    help = 'JSONify the relevant data for Twitterati members in the specified XLSX file and dump the results to the specified JSON file'

    def add_arguments(self, parser):
        # Positional arguments
        parser.add_argument('path_to_xlsx', type=str)
        parser.add_argument('path_to_json', type=str)

    def handle(self, *args, **options):
        json_data = self._jsonify_twitterati(options['path_to_xlsx'])
        with open(options['path_to_json'], 'w') as fp:
            json.dump(json_data, fp, indent=4)

    def _jsonify_twitterati(self, path_to_xlsx):
        self.stdout.write(self.style.NOTICE("JSONify Twitterati members defined in '{}'...".format(path_to_xlsx)))
        data = []
        categories = []
        members_by_category = {}
        members_with_no_category = []

        book = load_workbook(path_to_xlsx)
        number_of_sheets = len(book.worksheets)
        self.stdout.write(self.style.NOTICE("Found {} worksheet(s) named: {}'...".format(number_of_sheets, book.sheetnames)))
        if number_of_sheets > 1:
            self.stdout.write(self.style.WARNING("Assuming relevant data is found on the first worksheet; other worksheets will be ignored!"))
        elif number_of_sheets == 0:
            self.stdout.write(self.style.WARNING("Empty workbook!"))
            return data

        sheet = book.worksheets[0]
        for row in sheet.rows:
            values = [x.value.strip() for x in row]
            member = {
                'full_name': values[0],
                'twitter_handle': values[1],
                'biography': values[2],
            }
            category_name = values[3].lower()
            if len(category_name) > 0:
                if category_name not in categories:
                    categories.append(category_name)
                    members_by_category[category_name] = []
                members_by_category[category_name].append(member)
            else:
                members_with_no_category.append(member)

        categories.sort()
        total_members = 0
        if len(members_with_no_category) > 0:
            data.append({'name': 'NO CATEGORY', 'description': 'NO CATEGORY', 'members': members_with_no_category})
            member_count = len(members_with_no_category)
            total_members += member_count
            self.stdout.write(self.style.WARNING("Found {} Twitterati member(s) that do not belong to any category...".format(member_count)))
        for category_name in categories:
            sorted_members_by_category = sorted(members_by_category[category_name], key=itemgetter('full_name'))
            data.append({'category': category_name, 'description': category_name, 'members': sorted_members_by_category})
            member_count = len(sorted_members_by_category)
            total_members += member_count
            self.stdout.write(self.style.NOTICE("Found {} Twitterati member(s) that belong to the category '{}'...".format(member_count, category_name)))
        self.stdout.write(self.style.NOTICE("Found and JSONified {} Twitterati member(s) across {} category(ies).".format(total_members, len(categories))))
        return data
